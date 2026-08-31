"""
Measure SUL (urethral length) from the SUL_img3x per-video mask pairs.

Each video contributes one frame masked with the urethra (class 1) and one frame
masked with the catheter (class 4). Two ingredients give the length in mm:

  * the mirror plane of the urethra -- the line, at any angle, that best reflects
    the urethra mask onto itself. The urethra is usually tilted rather than
    upright, so the angle is searched too, not fixed to vertical. The urethra's
    extent along that line is the length in pixels.
  * the manually drawn arch over the catheter, from arches.json. The catheter is
    16/3 mm across, so the straight endpoint-to-endpoint distance of the arch
    (its chord, NOT the arc length -- the arches are drawn curved) sets the scale.

    SUL_mm = urethra_span_px * (16/3) / chord_px

python scripts/sul_measure.py --root "../transfer_atlas_mod/workspace/SUL_img3x"
"""
import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from PIL import Image
from scipy.fft import irfft, next_fast_len, rfft

URETHRA, CATHETER = 1, 4
CATHETER_MM = 16 / 3  # 16 Fr = 16/3 mm outer diameter

AXIS_BGR = (0, 255, 0)
URETHRA_BGR = (0, 0, 255)
ARCH_BGR = (255, 128, 0)


TIMESTAMP = re.compile(r"^\d{2}\.\d{2}\.\d{2}\.\d+$")


def video_id(stem):
    """Source video id: everything before the first timestamp.

    The timestamps that follow are lossless-cut artefacts (clip start/end, then
    frame offset), not part of the video's identity.
    """
    parts = stem.split("-")
    for i, part in enumerate(parts):
        if TIMESTAMP.match(part):
            return "-".join(parts[:i])
    return stem


def clip_key(stem):
    """Strip only the trailing frame offset, so both frames of a clip group."""
    return stem.rsplit("-", 1)[0]


def load_groups(root):
    """clip -> {class_id: stem} for the two annotated frames of that clip."""
    groups = defaultdict(dict)
    for path in sorted((root / "masks").glob("*.png")):
        mask = np.array(Image.open(path))
        for cls in np.unique(mask):
            if cls in (URETHRA, CATHETER):
                groups[clip_key(path.stem)][int(cls)] = path.stem
    return dict(groups)


def load_arches(root):
    """stem -> (left, right) endpoints. arches.json keys index the sorted images."""
    frames = json.loads((root / "arches.json").read_text())["frames"]
    images = sorted(p.stem for p in (root / "images").glob("*.png"))
    return {images[int(k)]: (np.array(v["left"]), np.array(v["right"]))
            for k, v in frames.items() if int(k) < len(images)}


def _scan_columns(mask):
    """Best vertical mirror line of `mask`: (axis x, overlap in pixels).

    overlap(q) = sum_r sum_c mask[r,c] * mask[r, q-c] counts, for a mirror centred
    at x = q/2, the pixels whose reflection is also set -- which is the row-wise
    self-convolution summed over rows, so one FFT gives every half-pixel
    candidate at once. Doing this per angle is what keeps the 2-D search cheap.
    """
    w = mask.shape[1]
    n = next_fast_len(2 * w - 1)  # zero-padding past 2w-1 leaves the result intact
    spec = rfft(mask.astype(np.float64), n=n, axis=1)
    overlaps = np.rint(irfft(spec * spec, n=n, axis=1).sum(0))[:2 * w - 1]

    # Flat maxima are common on near-symmetric shapes; sit in the middle of one.
    best = np.flatnonzero(overlaps == overlaps.max())
    return float(np.median(best)) / 2.0, int(overlaps.max())


def principal_axis(mask):
    """(tilt off vertical in degrees, elongation) of the mask's major axis."""
    ys, xs = np.nonzero(mask)
    spread = np.cov(np.stack([xs - xs.mean(), ys - ys.mean()]))
    val, vec = np.linalg.eigh(spread)
    major = vec[:, np.argmax(val)]
    tilt = np.degrees(np.arctan2(major[0], major[1]))
    return (tilt + 90) % 180 - 90, float(np.sqrt(max(val) / max(min(val), 1e-9)))


def mirror_axis(mask, max_tilt=40.0, coarse=1.0, fine=0.1):
    """Line that best mirrors `mask` onto itself, tilted up to `max_tilt` degrees.

    Symmetry is the sole criterion: every allowed tilt is scored and the most
    symmetric wins. The urethra is often rotated, so forcing the axis upright
    would score a tilted-but-symmetric mask as asymmetric and mismeasure its
    length. Rotating the mask by -angle turns the oblique search back into the
    vertical one above, and rotation is rigid so a span measured in the rotated
    frame is a true length.

    The tilt cap also keeps the search off the perpendicular solution: an
    elongated blob mirrors about its minor axis as well as its major one, and a
    wide-open search flips to measuring the urethra across instead of along.

    Returns the axis endpoints on the mask in original image coords, the tilt off
    vertical in degrees, the symmetry score (1.0 = perfect), and the span/filled
    pixel counts along the axis.
    """
    ys, xs = np.nonzero(mask)
    cx, cy = (xs.min() + xs.max()) / 2.0, (ys.min() + ys.max()) / 2.0
    side = int(np.hypot(ys.max() - ys.min(), xs.max() - xs.min())) + 8
    src = mask.astype(np.uint8)
    area = np.count_nonzero(mask)
    _, elongation = principal_axis(mask)  # reported for QC only, never steers the fit

    def rotated(angle):
        rot = cv2.getRotationMatrix2D((cx, cy), angle, 1.0)
        rot[0, 2] += side / 2.0 - cx
        rot[1, 2] += side / 2.0 - cy
        return rot, cv2.warpAffine(src, rot, (side, side), flags=cv2.INTER_NEAREST)

    def score(angle):
        _, canvas = rotated(angle)
        cols = np.flatnonzero(canvas.any(0))
        x, ov = _scan_columns(canvas[:, cols[0]:cols[-1] + 1])
        return ov, cols[0] + x

    # cv2 rotation by `a` tilts the resulting vertical axis by -a, so sweeping a
    # over +-max_tilt sweeps the axis over the same range either side of vertical.
    best_ov, best_angle, best_x = -1, 0.0, 0.0
    for stage in (np.arange(-max_tilt, max_tilt + coarse, coarse), None):
        if stage is None:  # refine around the coarse winner, staying inside the cap
            lo, hi = max(-max_tilt, best_angle - coarse), min(max_tilt, best_angle + coarse)
            stage = np.arange(lo, hi + fine, fine)
        for angle in stage:
            ov, x = score(angle)
            if ov > best_ov:
                best_ov, best_angle, best_x = ov, float(angle), x

    rot, canvas = rotated(best_angle)
    xi = int(round(best_x))
    col = np.flatnonzero(canvas[:, xi])
    neighbours = [np.ptp(np.flatnonzero(canvas[:, x])) + 1
                  for x in range(max(0, xi - 2), min(side, xi + 3))
                  if canvas[:, x].any()]

    inv = cv2.invertAffineTransform(rot)
    ends = np.array([[best_x, col[0]], [best_x, col[-1]]])
    p0, p1 = (inv[:, :2] @ ends.T).T + inv[:, 2]
    tilt = np.degrees(np.arctan2(p1[0] - p0[0], p1[1] - p0[1]))

    return {"p0": p0, "p1": p1, "tilt": float((tilt + 90) % 180 - 90),
            "score": best_ov / area, "elongation": elongation,
            "span": int(col[-1] - col[0] + 1), "filled": int(col.size),
            "span_med5": int(np.median(neighbours))}


def line_hits(mask, p0, p1):
    """True if the infinite line through p0,p1 crosses any set pixel of mask."""
    probe = np.zeros(mask.shape, np.uint8)
    d = p1 - p0
    far = 4000 * d / max(np.linalg.norm(d), 1e-6)
    cv2.line(probe, tuple(np.round(p0 - far).astype(int)),
             tuple(np.round(p1 + far).astype(int)), 1, 1)
    return bool((probe & mask.astype(np.uint8)).any())


def frame_bgr(images, stem):
    return cv2.cvtColor(np.array(Image.open(images / f"{stem}.png").convert("RGB")),
                        cv2.COLOR_RGB2BGR)


def write_xlsx(path, rows):
    """Same two columns as the result csv, with SUL as a real number not text."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SUL"
    ws.append(["Video ID", "SUL (mm)"])
    for r in rows:
        ws.append([r["video"], float(r["sul_mm"]) if r["sul_mm"] else None])

    for row in ws.iter_rows(max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.font = Font(name="Arial", size=11)
        row[1].number_format = "0.00"
        row[1].alignment = Alignment(horizontal="right")
    for cell in ws[1]:
        cell.font = Font(name="Arial", size=11, bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{ws.max_row}"
    ws.column_dimensions["A"].width = max(len(r["video"]) for r in rows) + 3
    ws.column_dimensions["B"].width = 12
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default="../transfer_atlas_mod/workspace/SUL_img3x")
    ap.add_argument("--csv", default=None, help="default: <root>/sul.csv")
    ap.add_argument("--max-tilt", type=float, default=40.0,
                    help="degrees the mirror axis may lean off vertical")
    ap.add_argument("--no-viz", action="store_true")
    args = ap.parse_args()

    root = Path(args.root)
    images, masks = root / "images", root / "masks"
    viz = root / "visualization"
    viz.mkdir(exist_ok=True)

    out = Path(args.csv) if args.csv else root / "sul.csv"
    full = out.with_name(f"{out.stem}_full.csv")
    xlsx = out.with_suffix(".xlsx")
    # Excel holds an exclusive lock on an open workbook; find out now rather than
    # after the measurement has already run.
    for path in (out, full, xlsx):
        try:
            open(path, "a").close()
        except PermissionError:
            raise SystemExit(f"[abort] {path.name} is open in another program -- close it")

    groups = load_groups(root)
    arches = load_arches(root)
    rows = []

    for clip, frames in sorted(groups.items()):
        video = video_id(clip)
        row = {"video": video, "urethra_frame": "", "catheter_frame": "",
               "axis_tilt_deg": "", "symmetry_score": "", "elongation": "",
               "urethra_span_px": "", "urethra_filled_px": "", "urethra_span_med5_px": "",
               "axis_x0_px": "", "axis_y0_px": "", "axis_x1_px": "", "axis_y1_px": "",
               "axis_crosses_catheter": "", "chord_px": "", "mm_per_px": "",
               "sul_mm": "", "status": ""}

        if URETHRA not in frames:
            row["catheter_frame"] = frames.get(CATHETER, "")
            row["status"] = "no urethra mask"
            rows.append(row)
            continue

        stem = frames[URETHRA]
        row["urethra_frame"] = stem
        u_mask = np.array(Image.open(masks / f"{stem}.png")) == URETHRA
        ax = mirror_axis(u_mask, max_tilt=args.max_tilt)
        span, p0, p1 = ax["span"], ax["p0"], ax["p1"]
        row.update(axis_tilt_deg=f"{ax['tilt']:.1f}", symmetry_score=f"{ax['score']:.3f}",
                   elongation=f"{ax['elongation']:.2f}",
                   urethra_span_px=span, urethra_filled_px=ax["filled"],
                   urethra_span_med5_px=ax["span_med5"],
                   axis_x0_px=f"{p0[0]:.1f}", axis_y0_px=f"{p0[1]:.1f}",
                   axis_x1_px=f"{p1[0]:.1f}", axis_y1_px=f"{p1[1]:.1f}")

        if not args.no_viz:
            img = frame_bgr(images, stem)
            contours, _ = cv2.findContours(u_mask.astype(np.uint8), cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)
            cv2.drawContours(img, contours, -1, URETHRA_BGR, 2)
            q0, q1 = (np.round(p).astype(int) for p in (p0, p1))
            unit = (q1 - q0) / max(np.linalg.norm(q1 - q0), 1e-6)
            perp = np.array([-unit[1], unit[0]]) * 25
            cv2.line(img, tuple(np.round(q0 - unit * 400).astype(int)),
                     tuple(np.round(q1 + unit * 400).astype(int)), (90, 90, 90), 1)
            cv2.line(img, tuple(q0), tuple(q1), AXIS_BGR, 2)
            for q in (q0, q1):
                cv2.line(img, tuple(np.round(q - perp).astype(int)),
                         tuple(np.round(q + perp).astype(int)), AXIS_BGR, 3)
            cv2.putText(img, f"urethra {span} px  tilt {ax['tilt']:+.1f} deg", (12, 42),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.1, AXIS_BGR, 2)
            cv2.imwrite(str(viz / f"{video}__urethra.png"), img)

        if CATHETER not in frames:
            row["status"] = "no catheter mask"
            rows.append(row)
            continue

        c_stem = frames[CATHETER]
        row["catheter_frame"] = c_stem
        c_mask = np.array(Image.open(masks / f"{c_stem}.png")) == CATHETER
        # The two frames are seconds apart, so the camera has usually drifted and
        # the urethra's axis need not land on the catheter. This only records how
        # far off the guide line is; the chord is valid wherever it was drawn.
        row["axis_crosses_catheter"] = "yes" if line_hits(c_mask, p0, p1) else "no"

        if c_stem not in arches:
            row["status"] = "no arch in arches.json"
            rows.append(row)
            continue

        left, right = arches[c_stem]
        chord = float(np.linalg.norm(left - right))  # endpoints only, not arc length
        if chord < 5:
            row["status"] = "degenerate arch"
            rows.append(row)
            continue

        scale = CATHETER_MM / chord
        row.update(chord_px=f"{chord:.1f}", mm_per_px=f"{scale:.5f}",
                   sul_mm=f"{span * scale:.2f}", status="ok")

        if not args.no_viz:
            img = frame_bgr(images, c_stem)
            contours, _ = cv2.findContours(c_mask.astype(np.uint8), cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)
            cv2.drawContours(img, contours, -1, (255, 255, 0), 2)
            p0, p1 = (tuple(np.round(p).astype(int)) for p in (left, right))
            cv2.line(img, p0, p1, ARCH_BGR, 3)
            for p in (p0, p1):
                cv2.circle(img, p, 8, ARCH_BGR, -1)
            cv2.putText(img, f"chord {chord:.1f} px = {CATHETER_MM:.2f} mm", (12, 42),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.1, ARCH_BGR, 2)
            cv2.imwrite(str(viz / f"{video}__catheter.png"), img)

        rows.append(row)

    # The result is just id + length; everything else is kept alongside for QC.
    with open(out, "w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["video_id", "sul_mm"])
        writer.writerows([r["video"], r["sul_mm"]] for r in rows)
    with open(full, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    write_xlsx(xlsx, rows)

    ok = [r for r in rows if r["status"] == "ok"]
    print(f"[done] {len(rows)} videos, {len(ok)} measured")
    print(f"  {out.name} / {out.with_suffix('.xlsx').name}  (video_id, sul_mm)")
    print(f"  {full.name}  (all diagnostics)")
    if ok:
        vals = np.array([float(r["sul_mm"]) for r in ok])
        print(f"  SUL mm: median {np.median(vals):.1f}  "
              f"IQR {np.percentile(vals, 25):.1f}-{np.percentile(vals, 75):.1f}  "
              f"range {vals.min():.1f}-{vals.max():.1f}")
    for r in rows:
        if r["status"] != "ok":
            print(f"  [skip] {r['video']}: {r['status']}")


if __name__ == "__main__":
    main()
